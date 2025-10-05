"""Testes de integração para interação entre spider e pipeline"""
import pytest
import os
from pathlib import Path
from scrapy.http import HtmlResponse, Request
from scrapy.utils.test import get_crawler
from scrapy import signals
from pydispatch import dispatcher
from lead_scraper.spiders.bing_maps_spider import BingMapsSpider
from lead_scraper.pipelines import ExcelExportPipeline
from lead_scraper.items import LeadScraperItem
import openpyxl


def test_spider_output_to_pipeline(mock_bing_html, tmp_path, monkeypatch):
    """
    Testa que itens fluem do spider para o pipeline corretamente.
    Verifica que itens gerados pelo spider são recebidos e processados adequadamente pelo pipeline.
    Requisitos: 4.1
    """
    # Configura spider
    crawler = get_crawler(BingMapsSpider)
    spider = BingMapsSpider.from_crawler(
        crawler,
        termo='academias',
        estado='RS',
        cidade='Porto Alegre',
        bairros=''
    )
    
    # Configura pipeline com diretório temporário
    temp_data_dir = tmp_path / 'data'
    temp_data_dir.mkdir(exist_ok=True)
    
    pipeline = ExcelExportPipeline()
    pipeline.open_spider(spider)
    # Sobrescreve results_folder após open_spider
    pipeline.results_folder = str(temp_data_dir)
    
    # Cria resposta simulada
    request = Request(url='https://www.bing.com/maps?q=academias')
    response = HtmlResponse(
        url='https://www.bing.com/maps?q=academias',
        request=request,
        body=mock_bing_html.encode('utf-8'),
        encoding='utf-8'
    )
    response.meta['bairro'] = None
    
    # Faz parse da resposta e coleta itens
    items = list(spider.parse(response))
    
    # Processa itens através do pipeline
    processed_items = []
    for item in items:
        processed = pipeline.process_item(item, spider)
        processed_items.append(processed)
    
    # Verifica que itens foram processados
    assert len(processed_items) > 0, "Pipeline deve processar pelo menos um item"
    
    # Verifica estrutura do item
    for item in processed_items:
        assert isinstance(item, LeadScraperItem)
        assert 'termo_busca' in item
        assert 'estado' in item
        assert 'cidade' in item
        assert 'nome' in item
        assert 'endereco' in item
    
    # Fecha pipeline e verifica criação de arquivo
    pipeline.close_spider(spider)
    
    # Verifica que arquivo Excel foi criado no diretório temporário
    excel_files = list(temp_data_dir.glob('leads_*.xlsx'))
    assert len(excel_files) == 1, f"Pipeline deve criar exatamente um arquivo Excel, encontrados {len(excel_files)}"
    
    # Verifica conteúdo Excel
    workbook = openpyxl.load_workbook(excel_files[0])
    sheet = workbook.active
    
    # Verifica cabeçalhos
    headers = [cell.value for cell in sheet[1]]
    assert headers == ["Termo", "Estado", "Cidade", "Bairro", "Nome", "Endereço", "Telefone", "Website"]
    
    # Verifica que linhas de dados correspondem aos itens processados
    data_rows = list(sheet.iter_rows(min_row=2, values_only=True))
    assert len(data_rows) == len(processed_items), "Excel deve conter todos os itens processados"


def test_multiple_items_through_pipeline(tmp_path):
    """
    Testa que múltiplos itens são processados corretamente através da cadeia completa.
    Verifica processamento em lote do spider ao pipeline até saída Excel.
    Requisitos: 4.1, 4.2
    """
    # Configura spider
    crawler = get_crawler(BingMapsSpider)
    spider = BingMapsSpider.from_crawler(
        crawler,
        termo='cafés',
        estado='SP',
        cidade='São Paulo',
        bairros=''
    )
    
    # Configura pipeline com diretório temporário
    temp_data_dir = tmp_path / 'data'
    temp_data_dir.mkdir(exist_ok=True)
    
    pipeline = ExcelExportPipeline()
    pipeline.open_spider(spider)
    # Sobrescreve results_folder após open_spider
    pipeline.results_folder = str(temp_data_dir)
    
    # Cria múltiplos itens de teste
    test_items = []
    for i in range(5):
        item = LeadScraperItem()
        item['termo_busca'] = 'cafés'
        item['estado'] = 'SP'
        item['cidade'] = 'São Paulo'
        item['bairro'] = f'Bairro {i+1}'
        item['nome'] = f'Café Test {i+1}'
        item['endereco'] = f'Rua Test {i+1}, 100'
        item['telefone'] = f'(11) 9999-{i:04d}'
        item['website'] = f'https://cafe{i+1}.com'
        test_items.append(item)
    
    # Processa todos os itens através do pipeline
    processed_count = 0
    for item in test_items:
        processed = pipeline.process_item(item, spider)
        assert processed is not None
        processed_count += 1
    
    assert processed_count == 5, "Todos os itens devem ser processados"
    
    # Fecha pipeline
    pipeline.close_spider(spider)
    
    # Verifica arquivo Excel apenas no diretório temporário
    excel_files = list(temp_data_dir.glob('leads_*.xlsx'))
    assert len(excel_files) == 1, f"Esperado 1 arquivo Excel, encontrados {len(excel_files)}"
    
    workbook = openpyxl.load_workbook(excel_files[0])
    sheet = workbook.active
    
    # Verifica que todos os itens estão no Excel (1 cabeçalho + 5 linhas de dados)
    assert sheet.max_row == 6, "Excel deve ter 1 linha de cabeçalho + 5 linhas de dados"
    
    # Verifica dados de cada item
    for i, row in enumerate(sheet.iter_rows(min_row=2, max_row=6, values_only=True), start=1):
        assert row[0] == 'cafés'  # termo_busca
        assert row[1] == 'SP'  # estado
        assert row[2] == 'São Paulo'  # cidade
        assert row[3] == f'Bairro {i}'  # bairro
        assert row[4] == f'Café Test {i}'  # nome
        assert row[5] == f'Rua Test {i}, 100'  # endereco
        assert row[6] == f'(11) 9999-{i-1:04d}'  # telefone
        assert row[7] == f'https://cafe{i}.com'  # website


def test_pipeline_receives_correct_item_structure(tmp_path):
    """
    Testa que pipeline recebe e valida estrutura correta do item.
    Verifica que itens estão em conformidade com a estrutura esperada do LeadScraperItem.
    Requisitos: 4.2
    """
    # Configura spider
    crawler = get_crawler(BingMapsSpider)
    spider = BingMapsSpider.from_crawler(
        crawler,
        termo='restaurantes',
        estado='RJ',
        cidade='Rio de Janeiro',
        bairros=''
    )
    
    # Configura pipeline com diretório temporário
    temp_data_dir = tmp_path / 'data'
    temp_data_dir.mkdir(exist_ok=True)
    
    pipeline = ExcelExportPipeline()
    pipeline.open_spider(spider)
    # Sobrescreve results_folder após open_spider
    pipeline.results_folder = str(temp_data_dir)
    
    # Cria item com todos os campos obrigatórios
    item = LeadScraperItem()
    item['termo_busca'] = 'restaurantes'
    item['estado'] = 'RJ'
    item['cidade'] = 'Rio de Janeiro'
    item['bairro'] = 'Copacabana'
    item['nome'] = 'Restaurante Teste'
    item['endereco'] = 'Av. Atlântica, 500'
    item['telefone'] = '(21) 3333-4444'
    item['website'] = 'https://restaurante.com'
    
    # Verifica estrutura do item antes do processamento
    required_fields = ['termo_busca', 'estado', 'cidade', 'bairro', 'nome', 'endereco', 'telefone', 'website']
    for field in required_fields:
        assert field in item, f"Item deve ter campo: {field}"
        assert item[field] is not None, f"Campo {field} não deve ser None"
    
    # Processa item através do pipeline
    processed = pipeline.process_item(item, spider)
    
    # Verifica que item processado mantém estrutura
    assert isinstance(processed, LeadScraperItem)
    for field in required_fields:
        assert field in processed
        assert processed[field] == item[field], f"Campo {field} deve permanecer inalterado após processamento"
    
    # Fecha pipeline
    pipeline.close_spider(spider)
    
    # Verifica que saída Excel tem estrutura correta apenas no diretório temporário
    excel_files = list(temp_data_dir.glob('leads_*.xlsx'))
    assert len(excel_files) == 1, f"Esperado 1 arquivo Excel, encontrados {len(excel_files)}"
    
    workbook = openpyxl.load_workbook(excel_files[0])
    sheet = workbook.active
    
    # Verifica que cabeçalhos correspondem aos campos do item
    headers = [cell.value for cell in sheet[1]]
    expected_headers = ["Termo", "Estado", "Cidade", "Bairro", "Nome", "Endereço", "Telefone", "Website"]
    assert headers == expected_headers, "Cabeçalhos Excel devem corresponder à estrutura do item"
    
    # Verifica linha de dados
    data_row = list(sheet.iter_rows(min_row=2, max_row=2, values_only=True))[0]
    assert data_row[0] == item['termo_busca']
    assert data_row[1] == item['estado']
    assert data_row[2] == item['cidade']
    assert data_row[3] == item['bairro']
    assert data_row[4] == item['nome']
    assert data_row[5] == item['endereco']
    assert data_row[6] == item['telefone']
    assert data_row[7] == item['website']
