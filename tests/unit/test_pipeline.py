import pytest
import os
import openpyxl
from datetime import datetime
from unittest.mock import Mock, patch
from lead_scraper.pipelines import ExcelExportPipeline
from lead_scraper.items import LeadScraperItem


class TestExcelExportPipeline:
    """Testes unitários para ExcelExportPipeline"""

    def test_pipeline_initialization(self, tmp_path, monkeypatch):
        """Verifica criação de workbook e sheet ao abrir spider"""
        # Simula o diretório de dados para usar tmp_path
        mock_pipeline_dir = tmp_path / 'lead_scraper' / 'lead_scraper'
        mock_pipeline_dir.mkdir(parents=True)
        
        with patch('lead_scraper.pipelines.os.path.dirname', return_value=str(mock_pipeline_dir)):
            pipeline = ExcelExportPipeline()
            mock_spider = Mock()
            
            pipeline.open_spider(mock_spider)
            
            # Verifica que workbook foi criado
            assert pipeline.workbook is not None
            assert isinstance(pipeline.workbook, openpyxl.Workbook)
            
            # Verifica que sheet foi criada e nomeada
            assert pipeline.sheet is not None
            assert pipeline.sheet.title == "Resultados"
            
            # Verifica que pasta de resultados foi definida
            assert pipeline.results_folder is not None
            assert os.path.exists(pipeline.results_folder)

    def test_pipeline_creates_headers(self, tmp_path):
        """Verifica cabeçalhos de coluna Excel corretos"""
        with patch('lead_scraper.pipelines.os.path.dirname', return_value=str(tmp_path)):
            pipeline = ExcelExportPipeline()
            mock_spider = Mock()
            
            pipeline.open_spider(mock_spider)
            
            # Obtém a primeira linha (cabeçalhos)
            headers = [cell.value for cell in pipeline.sheet[1]]
            
            expected_headers = [
                "Termo", "Estado", "Cidade", "Bairro", 
                "Nome", "Endereço", "Telefone", "Website"
            ]
            
            assert headers == expected_headers

    def test_pipeline_processes_single_item(self, tmp_path, sample_lead_item):
        """Verifica processamento de item único"""
        with patch('lead_scraper.pipelines.os.path.dirname', return_value=str(tmp_path)):
            pipeline = ExcelExportPipeline()
            mock_spider = Mock()
            
            pipeline.open_spider(mock_spider)
            result = pipeline.process_item(sample_lead_item, mock_spider)
            
            # Verifica que item é retornado (para encadeamento de pipeline)
            assert result == sample_lead_item
            
            # Verifica que dados foram escritos na planilha (linha 2, após cabeçalhos)
            row_data = [cell.value for cell in pipeline.sheet[2]]
            
            assert row_data[0] == sample_lead_item['termo_busca']
            assert row_data[1] == sample_lead_item['estado']
            assert row_data[2] == sample_lead_item['cidade']
            assert row_data[3] == sample_lead_item['bairro']
            assert row_data[4] == sample_lead_item['nome']
            assert row_data[5] == sample_lead_item['endereco']
            assert row_data[6] == sample_lead_item['telefone']
            assert row_data[7] == sample_lead_item['website']

    def test_pipeline_processes_multiple_items(self, tmp_path):
        """Verifica processamento em lote"""
        with patch('lead_scraper.pipelines.os.path.dirname', return_value=str(tmp_path)):
            pipeline = ExcelExportPipeline()
            mock_spider = Mock()
            
            pipeline.open_spider(mock_spider)
            
            # Cria múltiplos itens
            items = []
            for i in range(3):
                item = LeadScraperItem()
                item['termo_busca'] = f'termo_{i}'
                item['estado'] = f'Estado_{i}'
                item['cidade'] = f'Cidade_{i}'
                item['bairro'] = f'Bairro_{i}'
                item['nome'] = f'Nome_{i}'
                item['endereco'] = f'Endereco_{i}'
                item['telefone'] = f'Telefone_{i}'
                item['website'] = f'https://example{i}.com'
                items.append(item)
                pipeline.process_item(item, mock_spider)
            
            # Verifica que todos os itens foram escritos (linhas 2, 3, 4 após linha de cabeçalho 1)
            assert pipeline.sheet.max_row == 4  # 1 cabeçalho + 3 itens
            
            # Verifica dados de cada item
            for i, item in enumerate(items, start=2):
                row_data = [cell.value for cell in pipeline.sheet[i]]
                assert row_data[0] == items[i-2]['termo_busca']
                assert row_data[4] == items[i-2]['nome']

    def test_pipeline_handles_unicode_characters(self, tmp_path):
        """Verifica codificação de caracteres portugueses"""
        with patch('lead_scraper.pipelines.os.path.dirname', return_value=str(tmp_path)):
            pipeline = ExcelExportPipeline()
            mock_spider = Mock()
            
            pipeline.open_spider(mock_spider)
            
            # Cria item com caracteres portugueses
            item = LeadScraperItem()
            item['termo_busca'] = 'açougue'
            item['estado'] = 'São Paulo'
            item['cidade'] = 'São José dos Campos'
            item['bairro'] = 'Jardim São Dimas'
            item['nome'] = 'Açougue do João'
            item['endereco'] = 'Rua da Conceição, 123'
            item['telefone'] = '(12) 3456-7890'
            item['website'] = 'https://açougue.com.br'
            
            pipeline.process_item(item, mock_spider)
            
            # Verifica que caracteres portugueses são preservados
            row_data = [cell.value for cell in pipeline.sheet[2]]
            assert row_data[0] == 'açougue'
            assert row_data[1] == 'São Paulo'
            assert row_data[2] == 'São José dos Campos'
            assert row_data[3] == 'Jardim São Dimas'
            assert row_data[4] == 'Açougue do João'
            assert row_data[5] == 'Rua da Conceição, 123'

    def test_pipeline_creates_output_file(self, tmp_path):
        """Verifica criação de arquivo com formato de timestamp correto"""
        with patch('lead_scraper.pipelines.os.path.dirname', return_value=str(tmp_path)):
            pipeline = ExcelExportPipeline()
            mock_spider = Mock()
            
            pipeline.open_spider(mock_spider)
            
            # Simula datetime para controlar timestamp
            mock_datetime = datetime(2025, 10, 4, 14, 30, 45)
            with patch('lead_scraper.pipelines.datetime') as mock_dt:
                mock_dt.datetime.now.return_value = mock_datetime
                
                pipeline.close_spider(mock_spider)
                
                # Verifica que arquivo foi criado com formato de timestamp correto
                expected_filename = "leads_20251004_143045.xlsx"
                expected_filepath = os.path.join(pipeline.results_folder, expected_filename)
                
                assert os.path.exists(expected_filepath)
                
                # Verifica que é um arquivo Excel válido
                workbook = openpyxl.load_workbook(expected_filepath)
                assert workbook is not None
                workbook.close()

    def test_pipeline_creates_data_directory(self, tmp_path):
        """Verifica criação do diretório data se ausente"""
        # Cria um caminho que ainda não existe
        mock_pipeline_dir = tmp_path / 'lead_scraper' / 'lead_scraper'
        mock_pipeline_dir.mkdir(parents=True)
        
        # Garante que diretório data não existe
        data_dir = tmp_path / 'data'
        assert not data_dir.exists()
        
        with patch('lead_scraper.pipelines.os.path.dirname', return_value=str(mock_pipeline_dir)):
            with patch('lead_scraper.pipelines.os.path.abspath', return_value=str(data_dir)):
                pipeline = ExcelExportPipeline()
                mock_spider = Mock()
                
                pipeline.open_spider(mock_spider)
                
                # Verifica que diretório data foi criado
                assert os.path.exists(pipeline.results_folder)
                assert pipeline.results_folder == str(data_dir)

    def test_excel_file_structure(self, tmp_path, sample_lead_item):
        """Verifica que arquivo Excel pode ser aberto e lido"""
        with patch('lead_scraper.pipelines.os.path.dirname', return_value=str(tmp_path)):
            pipeline = ExcelExportPipeline()
            mock_spider = Mock()
            
            pipeline.open_spider(mock_spider)
            pipeline.process_item(sample_lead_item, mock_spider)
            
            # Salva o arquivo
            mock_datetime = datetime(2025, 10, 4, 15, 0, 0)
            with patch('lead_scraper.pipelines.datetime') as mock_dt:
                mock_dt.datetime.now.return_value = mock_datetime
                pipeline.close_spider(mock_spider)
            
            # Abre e verifica a estrutura do arquivo
            filepath = os.path.join(pipeline.results_folder, "leads_20251004_150000.xlsx")
            workbook = openpyxl.load_workbook(filepath)
            
            # Verifica que sheet existe
            assert "Resultados" in workbook.sheetnames
            sheet = workbook["Resultados"]
            
            # Verifica cabeçalhos
            headers = [cell.value for cell in sheet[1]]
            assert headers == [
                "Termo", "Estado", "Cidade", "Bairro",
                "Nome", "Endereço", "Telefone", "Website"
            ]
            
            # Verifica linha de dados
            data_row = [cell.value for cell in sheet[2]]
            assert data_row[0] == 'academias'
            assert data_row[4] == 'Academia Fitness Plus'
            assert data_row[6] == '(51) 3333-4444'
            
            workbook.close()
